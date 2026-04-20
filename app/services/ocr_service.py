import re
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


class OCRService:
    """Service voor het extraheren van tekst uit documenten (PDF, afbeelding, CSV, XLSX)."""

    async def extract_text(self, file_content: bytes, file_type: str) -> str:
        """Extraheer tekst uit een bestand op basis van het type."""
        ft = (file_type or "").lower().lstrip(".")
        if ft == "pdf":
            return await self.extract_from_pdf(file_content)
        if ft in ("jpg", "jpeg", "png"):
            return await self.extract_from_image(file_content)
        if ft == "csv":
            return self._decode_text(file_content)
        if ft in ("xlsx", "xls"):
            return self.extract_from_xlsx(file_content)
        logger.warning(f"OCRService: onbekend bestandstype '{file_type}'")
        return ""

    @staticmethod
    def _decode_text(content: bytes) -> str:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return content.decode(enc)
            except UnicodeDecodeError:
                continue
        return content.decode("utf-8", errors="replace")

    async def extract_from_pdf(self, content: bytes) -> str:
        """Extraheer tekst uit een PDF bestand.

        Probeert eerst pdfplumber (snelle tekst-extractie voor digitale PDFs).
        Valt terug op OCR via pdf2image + pytesseract voor gescande PDFs.
        """
        text = ""
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                pages_text = []
                for page in pdf.pages:
                    t = page.extract_text() or ""
                    if t.strip():
                        pages_text.append(t)
                text = "\n\n".join(pages_text).strip()
        except ImportError:
            logger.warning("pdfplumber niet geïnstalleerd — val terug op OCR")
        except Exception as e:
            logger.warning(f"pdfplumber extractie mislukt: {e}")

        if text:
            return text

        # Fallback: render PDF naar afbeeldingen en OCR elke pagina
        try:
            from pdf2image import convert_from_bytes
            import pytesseract

            images = convert_from_bytes(content, dpi=200)
            ocr_pages = []
            for img in images:
                try:
                    ocr_pages.append(pytesseract.image_to_string(img, lang="nld+eng"))
                except Exception:
                    ocr_pages.append(pytesseract.image_to_string(img))
            return "\n\n".join(p for p in ocr_pages if p.strip()).strip()
        except ImportError:
            logger.warning("pdf2image/pytesseract niet beschikbaar voor OCR-fallback")
            return ""
        except Exception as e:
            logger.error(f"OCR-fallback mislukt: {e}")
            return ""

    async def extract_from_image(self, content: bytes) -> str:
        """Extraheer tekst uit een afbeelding via OCR."""
        try:
            import pytesseract
            from PIL import Image

            image = Image.open(io.BytesIO(content))
            try:
                return pytesseract.image_to_string(image, lang="nld+eng")
            except Exception:
                return pytesseract.image_to_string(image)
        except ImportError:
            logger.warning("pytesseract/PIL niet geïnstalleerd")
            return ""
        except Exception as e:
            logger.error(f"Image OCR mislukt: {e}")
            return ""

    def extract_from_xlsx(self, content: bytes) -> str:
        """Converteer XLSX naar platte tekst (één regel per rij, cellen tab-gescheiden)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl niet geïnstalleerd")
            return ""

        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            logger.error(f"Kan XLSX niet openen: {e}")
            return ""

        sheets_text = []
        for sheet in wb.worksheets:
            lines = [f"## Sheet: {sheet.title}"]
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells))
            sheets_text.append("\n".join(lines))
        return "\n\n".join(sheets_text).strip()

    def parse_invoice_data(self, text: str) -> dict[str, Any]:
        """Parse factuurgegevens uit geëxtraheerde tekst."""
        data: dict[str, Any] = {
            "date": None,
            "amount": None,
            "btw_amount": None,
            "vendor": None,
            "invoice_number": None,
        }

        # Datumpatronen: DD-MM-YYYY, DD/MM/YYYY
        date_pattern = r"(\d{2}[-/]\d{2}[-/]\d{4})"
        date_match = re.search(date_pattern, text)
        if date_match:
            data["date"] = date_match.group(1)

        # Bedragpatronen: €1.234,56 of EUR 1234.56
        amount_patterns = [
            r"[€]\s*([\d.,]+)",
            r"EUR\s*([\d.,]+)",
            r"totaal[:\s]*([\d.,]+)",
        ]
        for pattern in amount_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                amount_str = match.group(1).replace(".", "").replace(",", ".")
                try:
                    data["amount"] = float(amount_str)
                except ValueError:
                    pass
                break

        # BTW bedrag
        btw_patterns = [r"BTW[:\s]*([\d.,]+)", r"btw[:\s]*[€]?\s*([\d.,]+)"]
        for pattern in btw_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                btw_str = match.group(1).replace(".", "").replace(",", ".")
                try:
                    data["btw_amount"] = float(btw_str)
                except ValueError:
                    pass
                break

        # Factuurnummer
        inv_patterns = [
            r"factuurnummer[:\s]*([A-Za-z0-9-]+)",
            r"factuur\s*nr[.:\s]*([A-Za-z0-9-]+)",
            r"invoice[:\s]*([A-Za-z0-9-]+)",
        ]
        for pattern in inv_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                data["invoice_number"] = match.group(1).strip()
                break

        return data
