"""CSV import pipeline — accepts CSV uploads from SnelStart, Moneybird, WeFact, or generic format.

Uses AI to classify rows and map columns, then creates native Invoice/Expense/BankTransaction records.
"""
import uuid
import csv
import io
import logging
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.models.user import User
from app.models.invoice import Invoice
from app.models.expense import Expense
from app.utils.auth import get_current_user

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/csv-import", tags=["CSV Import"])


def _safe_float(raw) -> float:
    if not raw:
        return 0.0
    try:
        # Handle Dutch number format: 1.234,56 → 1234.56
        s = str(raw).strip().replace('€', '').replace(' ', '')
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            s = s.replace(',', '.')
        return float(s)
    except (TypeError, ValueError):
        return 0.0


def _safe_date(raw) -> date:
    if not raw:
        return date.today()
    s = str(raw).strip()
    # Try common Dutch date formats
    for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d.%m.%Y']:
        try:
            from datetime import datetime
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return date.today()


def _detect_format(headers: list[str]) -> str:
    """Detect CSV format from headers."""
    h_lower = [h.lower().strip() for h in headers]
    h_joined = ' '.join(h_lower)

    if 'factuurnummer' in h_joined and ('relatie' in h_joined or 'naam' in h_joined):
        return 'snelstart'
    if 'invoicecode' in h_joined or 'debtorcode' in h_joined:
        return 'wefact'
    if 'invoice_id' in h_joined or 'contact_id' in h_joined:
        return 'moneybird'
    if 'order_number' in h_joined or 'lineitem' in h_joined:
        return 'shopify'
    return 'generic'


def _guess_type(row: dict, headers_lower: list[str]) -> str:
    """Guess if a row is income (verkoopfactuur) or expense (inkoopfactuur/uitgave)."""
    for key in ['type', 'soort', 'categorie', 'category']:
        val = str(row.get(key, '')).lower()
        if any(w in val for w in ['inkoop', 'kosten', 'uitgave', 'expense', 'purchase', 'credit']):
            return 'expense'
        if any(w in val for w in ['verkoop', 'omzet', 'revenue', 'income', 'debit']):
            return 'invoice'

    # Check amount sign
    for key in ['bedrag', 'amount', 'total', 'bedrag incl. btw', 'bedrag excl. btw', 'totaal']:
        for h in row:
            if key in h.lower():
                val = _safe_float(row[h])
                if val < 0:
                    return 'expense'
                if val > 0:
                    return 'invoice'
    return 'invoice'


def _guess_btw(row: dict) -> float:
    """Extract BTW rate from row."""
    for key in row:
        kl = key.lower()
        if 'btw%' in kl or 'btw tarief' in kl or 'tax_rate' in kl or 'taxpercentage' in kl or 'vat' in kl:
            val = _safe_float(row[key])
            if 0 < val <= 100:
                return val
    # Try to compute from excl/incl
    excl = incl = 0
    for key in row:
        kl = key.lower()
        if 'excl' in kl:
            excl = _safe_float(row[key])
        if 'incl' in kl:
            incl = _safe_float(row[key])
    if excl > 0 and incl > excl:
        rate = ((incl - excl) / excl) * 100
        if abs(rate - 21) < 3:
            return 21
        if abs(rate - 9) < 3:
            return 9
        return round(rate)
    return 21


def _find_field(row: dict, candidates: list[str], default: str = '') -> str:
    """Find a field value by trying multiple column name candidates."""
    for key in row:
        kl = key.lower().strip()
        for c in candidates:
            if c in kl:
                return str(row[key]).strip()
    return default


def _parse_row(row: dict, fmt: str) -> dict:
    """Parse a single CSV row into a normalized item."""
    number = _find_field(row, ['factuurnummer', 'invoice', 'nummer', 'number', 'code', 'referentie', 'ref'])
    client = _find_field(row, ['relatie', 'klant', 'naam', 'customer', 'client', 'contact', 'company', 'debtor', 'crediteur'])
    desc = _find_field(row, ['omschrijving', 'description', 'memo', 'notitie', 'toelichting', 'product'])
    date_str = _find_field(row, ['datum', 'date', 'factuurdatum', 'invoice_date', 'boekdatum'])
    amount = 0.0

    # Try multiple amount fields
    for key in row:
        kl = key.lower()
        if any(w in kl for w in ['bedrag incl', 'amount_incl', 'totaal', 'total', 'bedrag']):
            v = _safe_float(row[key])
            if v != 0:
                amount = v
                break
    if amount == 0:
        for key in row:
            kl = key.lower()
            if any(w in kl for w in ['bedrag', 'amount', 'prijs', 'price', 'waarde']):
                v = _safe_float(row[key])
                if v != 0:
                    amount = v
                    break

    category = _find_field(row, ['categorie', 'category', 'grootboek', 'ledger', 'account', 'kostenplaats'])
    row_type = _guess_type(row, [k.lower() for k in row.keys()])
    btw_rate = _guess_btw(row)

    return {
        "number": number or f"CSV-{uuid.uuid4().hex[:6].upper()}",
        "client": client,
        "description": desc or f"CSV import: {number}",
        "date": date_str,
        "amount": abs(amount),
        "btw_rate": btw_rate,
        "category": category or "Overig",
        "type": row_type,
        "raw": row,
    }


@router.post("/parse")
async def parse_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Parse a CSV or XLSX file and return structured items for review.

    Supports: SnelStart, Moneybird, WeFact, Shopify, and generic CSV/XLSX formats.
    Auto-detects format from headers. Returns items with suggested type,
    amount, BTW, category, etc. for the user to review before accepting.
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Alleen CSV of XLSX bestanden worden ondersteund")

    content = await file.read()

    if filename.endswith(('.xlsx', '.xls')):
        try:
            headers, rows = _read_xlsx_rows(content)
        except ImportError:
            raise HTTPException(status_code=500, detail="XLSX-ondersteuning is niet geïnstalleerd op de server")
        except Exception as e:
            logger.warning(f"XLSX parse error: {e}")
            raise HTTPException(status_code=400, detail=f"Kon XLSX niet lezen: {e}")
    else:
        # Try UTF-8 first, then Latin-1 (common in Dutch exports)
        try:
            text = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = content.decode('latin-1')

        # Detect delimiter
        first_line = text.split('\n')[0]
        delimiter = ';' if ';' in first_line else ','

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = list(reader.fieldnames or [])
        rows = list(reader)

    fmt = _detect_format(headers)

    items = []
    for i, row in enumerate(rows):
        if i >= 500:  # Safety limit
            break
        try:
            parsed = _parse_row(row, fmt)
            if parsed["amount"] > 0 or parsed["client"]:
                items.append({
                    "id": i + 1,
                    "document": file.filename,
                    **parsed,
                    "confidence": 0.85 if parsed["client"] else 0.5,
                    "accepted": False,
                })
        except Exception as e:
            logger.warning(f"Row {i} parse error: {e}")

    return {
        "format": fmt,
        "headers": headers,
        "total_rows": len(items),
        "items": items,
    }


def _read_xlsx_rows(content: bytes) -> tuple[list[str], list[dict]]:
    """Read an XLSX file and return (headers, [row-dict, ...]) like csv.DictReader."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = wb.active
    if sheet is None:
        return [], []

    iterator = sheet.iter_rows(values_only=True)
    try:
        first = next(iterator)
    except StopIteration:
        return [], []

    headers = [("" if v is None else str(v)).strip() for v in first]
    rows: list[dict] = []
    for row in iterator:
        if row is None:
            continue
        values = list(row) + [None] * (len(headers) - len(row))
        rec = {}
        for h, v in zip(headers, values):
            if not h:
                continue
            rec[h] = "" if v is None else v
        if any(str(v).strip() for v in rec.values()):
            rows.append(rec)
    return headers, rows


@router.post("/accept")
async def accept_items(
    data: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Accept parsed CSV items and create native Invoice/Expense records."""
    items = data.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="Geen items om te importeren")

    company_id = current_user.company_id
    if not company_id:
        raise HTTPException(status_code=400, detail="Geen bedrijf gekoppeld")

    invoices_created = 0
    expenses_created = 0

    for item in items:
        if not item.get("accepted"):
            continue

        d = _safe_date(item.get("date"))
        amount = abs(_safe_float(item.get("amount")))
        btw = _safe_float(item.get("btw_rate", 21))
        if abs(btw - 21) < 3:
            btw = 21
        elif abs(btw - 9) < 3:
            btw = 9
        elif abs(btw) < 2:
            btw = 0

        if item.get("type") == "expense":
            db.add(Expense(
                id=uuid.uuid4(),
                company_id=company_id,
                date=d,
                description=item.get("description") or item.get("client") or "CSV import",
                category=item.get("category") or "Overig",
                amount=amount,
                btw_rate=btw,
                status="geboekt",
                source="csv",
                source_id=f"csv-{item.get('number', '')}",
            ))
            expenses_created += 1
        else:
            db.add(Invoice(
                id=uuid.uuid4(),
                company_id=company_id,
                number=item.get("number") or f"CSV-{uuid.uuid4().hex[:6]}",
                client=item.get("client") or "Onbekend",
                date=d,
                due_date=d,
                lines=[{
                    "desc": item.get("description") or item.get("number") or "CSV import",
                    "qty": 1,
                    "price": amount,
                    "btwRate": btw,
                }],
                status="betaald",
                source="csv",
                source_id=f"csv-{item.get('number', '')}",
            ))
            invoices_created += 1

    await db.flush()
    return {
        "invoices_created": invoices_created,
        "expenses_created": expenses_created,
        "total_accepted": invoices_created + expenses_created,
    }
